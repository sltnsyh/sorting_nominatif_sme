import pandas as pd
from openpyxl import load_workbook
import os

def auto_pivot_excel():
    print("Pivot Excel Data by 'KODE CABANG'")
    # Get input and output file paths from the user
    input_file = input("Enter the full path to the input Excel file: ").strip()
    output_file = input("Enter the desired output Excel file path: ").strip()

    # Validate input file
    if not os.path.isfile(input_file):
        print("Error: Input file does not exist.")
        return

    try:
        # Load the Excel file into a pandas DataFrame
        df = pd.read_excel(input_file)

        # Ensure `KODE CABANG` is treated as a string (for consistent grouping)
        df["KODE CABANG"] = df["KODE CABANG"].astype(str)

        # Create a dictionary to hold data for each unique `KODE CABANG`
        grouped_data = {kode_cabang: group for kode_cabang, group in df.groupby("KODE CABANG")}

        # Write each group's data to a new sheet in the output Excel file
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for kode_cabang, group_data in grouped_data.items():
                # Use the `KODE CABANG` as the sheet name
                sheet_name = f"KODE_{kode_cabang[:30]}"  # Limit sheet name to 30 characters
                group_data.to_excel(writer, sheet_name=sheet_name, index=False)

        # Auto-adjust column widths
        wb = load_workbook(output_file)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for column in sheet.columns:
                max_length = 0
                col_letter = column[0].column_letter  # Get the column letter
                for cell in column:
                    try:
                        if cell.value:  # Calculate the maximum width of each column
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                adjusted_width = max_length + 2  # Add some padding
                sheet.column_dimensions[col_letter].width = adjusted_width

        # Save the workbook with adjusted column widths
        wb.save(output_file)

        print(f"Data has been pivoted and saved with adjusted headers to {output_file}")
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    auto_pivot_excel()
